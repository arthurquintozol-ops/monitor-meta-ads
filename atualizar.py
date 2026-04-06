"""
Script CLI para atualizar a planilha de monitoramento de anúncios da Meta.
Lê os links da coluna "Link da biblioteca" na planilha existente,
consulta cada um, e grava os resultados numa nova coluna com a data de hoje.

Uso: python atualizar.py [caminho_planilha.xlsx]
Se nenhum caminho for informado, usa "monitor_meta_ads.xlsx" na pasta atual.
"""

import re
import sys
from pathlib import Path
from urllib.parse import urlparse, parse_qs
from datetime import datetime

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


ARQUIVO_EXCEL_PADRAO = "monitor_meta_ads.xlsx"
ABA_EXCEL = "Historico"

FILL_SUBIU = PatternFill(fill_type="solid", fgColor="C6EFCE")
FILL_CAIU = PatternFill(fill_type="solid", fgColor="FFC7CE")
FILL_IGUAL = PatternFill(fill_type="solid", fgColor="FFEB9C")

FONT_SUBIU = Font(color="006100")
FONT_CAIU = Font(color="9C0006")
FONT_IGUAL = Font(color="9C6500")
FONT_NORMAL = Font(color="000000")
FONT_HEADER = Font(bold=True, color="000000")
FILL_HEADER = PatternFill(fill_type="solid", fgColor="D9E2F3")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


# ── JavaScript que extrai tudo numa única chamada ──

JS_EXTRAIR_DADOS = """
() => {
    let nomePagina = "Não identificado";
    const body = document.body ? document.body.innerText : "";
    const linhas = body.split("\\n").map(l => l.trim()).filter(l => l.length > 0);
    const ignorar = new Set([
        "Biblioteca de Anúncios da Meta", "Biblioteca de Anúncios",
        "Relatório da Biblioteca de Anúncios", "API da Biblioteca de Anúncios",
        "Conteúdo de marca", "Anúncios", "Sobre", "Entrar", "Status do sistema"
    ]);

    for (let i = 0; i < Math.min(linhas.length, 120); i++) {
        if ((linhas[i] === "Anúncios" || linhas[i] === "Sobre") && i > 0) {
            const candidato = linhas[i - 1].trim();
            if (candidato && candidato.length < 100 && !ignorar.has(candidato)) {
                nomePagina = candidato;
                break;
            }
        }
    }

    if (nomePagina === "Não identificado") {
        for (const tag of ["h1", "h2", "h3", "strong"]) {
            const els = document.querySelectorAll(tag);
            for (const el of els) {
                const t = el.innerText.trim().replace(/\\s+/g, " ");
                if (t && t.length < 100 && !ignorar.has(t)) {
                    nomePagina = t;
                    break;
                }
            }
            if (nomePagina !== "Não identificado") break;
        }
    }

    const bodyNorm = body.replace(/\\s+/g, " ");
    const regexes = [
        /~\\s*([\\d.,]+)\\s*resultados?/i,
        /aproximadamente\\s*([\\d.,]+)\\s*resultados?/i,
        /([\\d.,]+)\\s*resultados?/i
    ];

    let total = null;
    let textoOriginal = null;

    for (const rx of regexes) {
        const m = bodyNorm.match(rx);
        if (m) {
            textoOriginal = m[0];
            const numLimpo = m[1].replace(/\\D/g, "");
            if (numLimpo) {
                total = parseInt(numLimpo, 10);
                break;
            }
        }
    }

    return { nomePagina, total, textoOriginal };
}
"""


# ── Funções de extração ──

def extrair_page_id_da_url(url: str) -> str:
    try:
        query = parse_qs(urlparse(url).query)
        return str(query.get("view_all_page_id", [""])[0]).strip()
    except Exception:
        return ""


def gerar_prefixo_debug(url: str, indice: int) -> str:
    try:
        query = parse_qs(urlparse(url).query)
        page_id = query.get("view_all_page_id", ["sem_id"])[0]
        page_id = re.sub(r"[^\w\-]", "_", page_id)
        return f"{indice:03d}_{page_id}"
    except Exception:
        return f"{indice:03d}_sem_id"


def extrair_com_js(page, url: str, page_id: str) -> dict:
    pagina = "Não identificado"

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=60000)

        try:
            page.locator("text=resultado").first.wait_for(timeout=12000)
        except Exception:
            try:
                page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass

        dados = page.evaluate(JS_EXTRAIR_DADOS)

        pagina = dados.get("nomePagina") or "Não identificado"
        total = dados.get("total")
        texto_original = dados.get("textoOriginal")

        if total is not None:
            return {
                "status": "ok",
                "url": url,
                "pagina": pagina,
                "page_id": page_id,
                "total_anuncios_ativos": total,
                "texto_original": texto_original or "",
            }

        return {
            "status": "erro",
            "url": url,
            "pagina": pagina,
            "page_id": page_id,
            "total_anuncios_ativos": 0,
            "mensagem": "Texto de resultados não encontrado.",
        }

    except PlaywrightTimeoutError:
        return {
            "status": "erro",
            "url": url,
            "pagina": pagina,
            "page_id": page_id,
            "total_anuncios_ativos": 0,
            "mensagem": "Timeout ao carregar página.",
        }

    except Exception as e:
        return {
            "status": "erro",
            "url": url,
            "pagina": pagina,
            "page_id": page_id,
            "total_anuncios_ativos": 0,
            "mensagem": f"Erro: {str(e)}",
        }


# ── Funções de leitura/escrita do Excel ──

def ler_links_da_planilha(caminho: str) -> list[str]:
    """Lê todos os links da coluna 'Link da biblioteca' na planilha existente."""
    arquivo = Path(caminho)
    if not arquivo.exists():
        print(f"ERRO: Planilha '{caminho}' não encontrada.")
        sys.exit(1)

    wb = load_workbook(arquivo, read_only=True)
    if ABA_EXCEL not in wb.sheetnames:
        print(f"ERRO: Aba '{ABA_EXCEL}' não encontrada na planilha.")
        wb.close()
        sys.exit(1)

    ws = wb[ABA_EXCEL]

    col_link = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Link da biblioteca":
            col_link = col
            break

    if col_link is None:
        print("ERRO: Coluna 'Link da biblioteca' não encontrada no cabeçalho.")
        wb.close()
        sys.exit(1)

    links = []
    vistos = set()
    for row in range(2, ws.max_row + 1):
        valor = ws.cell(row=row, column=col_link).value
        if valor:
            url = str(valor).strip()
            if "facebook.com/ads/library" in url and url not in vistos:
                vistos.add(url)
                links.append(url)

    wb.close()
    return links


def inicializar_estrutura(ws):
    ws.cell(row=1, column=1, value="Nome da página")
    ws.cell(row=1, column=2, value="Link da biblioteca")


def encontrar_coluna_link(ws) -> int | None:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Link da biblioteca":
            return col
    return None


def obter_ou_criar_coluna_data(ws, data_str: str) -> int:
    col_link = encontrar_coluna_link(ws)

    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == data_str:
            return col

    if col_link is None:
        nova_coluna = ws.max_column + 1
        ws.cell(row=1, column=nova_coluna, value=data_str)
        return nova_coluna

    ws.insert_cols(col_link)
    ws.cell(row=1, column=col_link, value=data_str)
    return col_link


def obter_ou_reposicionar_coluna_link(ws) -> int:
    col_link = encontrar_coluna_link(ws)

    if col_link is None:
        col_link = ws.max_column + 1
        ws.cell(row=1, column=col_link, value="Link da biblioteca")
        return col_link

    if col_link != ws.max_column:
        garantir_link_na_ultima_coluna(ws)
        return ws.max_column

    return col_link


def garantir_link_na_ultima_coluna(ws):
    col_link = encontrar_coluna_link(ws)
    if col_link is None:
        ws.cell(row=1, column=ws.max_column + 1, value="Link da biblioteca")
        return

    if col_link == ws.max_column:
        return

    ultima_coluna = ws.max_column + 1
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=ultima_coluna, value=ws.cell(row=row, column=col_link).value)
    ws.delete_cols(col_link, 1)


def criar_chave_registro(page_id: str, pagina: str, link: str) -> str:
    page_id = (page_id or "").strip().lower()
    pagina = (pagina or "").strip().lower()
    link = (link or "").strip().lower()
    if page_id:
        return f"id::{page_id}"
    return f"nome::{pagina}||link::{link}"


def montar_indice_paginas(ws) -> dict:
    indice = {}
    col_link = encontrar_coluna_link(ws)

    for row in range(2, ws.max_row + 1):
        pagina = ws.cell(row=row, column=1).value
        link = ws.cell(row=row, column=col_link).value if col_link else ""

        pagina = str(pagina).strip() if pagina else ""
        link = str(link).strip() if link else ""

        page_id = extrair_page_id_da_url(link)
        chave = criar_chave_registro(page_id, pagina, link)

        if pagina or link:
            indice[chave] = row

    return indice


def encontrar_coluna_data_anterior(ws, col_data_atual: int) -> int | None:
    for col in range(col_data_atual - 1, 1, -1):
        titulo = ws.cell(row=1, column=col).value
        if titulo and titulo != "Link da biblioteca":
            return col
    return None


def converter_para_numero(valor):
    if valor is None or valor == "":
        return None
    try:
        return float(valor)
    except Exception:
        try:
            texto = str(valor).strip().replace(".", "").replace(",", ".")
            return float(texto)
        except Exception:
            return None


def salvar_resultados_no_excel(resultados: list[dict], caminho_arquivo: str):
    data_hoje = datetime.now().strftime("%d/%m/%Y")
    arquivo = Path(caminho_arquivo)
    arquivo.parent.mkdir(parents=True, exist_ok=True)

    if arquivo.exists():
        wb = load_workbook(arquivo)
        if ABA_EXCEL in wb.sheetnames:
            ws = wb[ABA_EXCEL]
        else:
            ws = wb.create_sheet(ABA_EXCEL)
            inicializar_estrutura(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = ABA_EXCEL
        inicializar_estrutura(ws)

    col_data = obter_ou_criar_coluna_data(ws, data_hoje)
    col_link = obter_ou_reposicionar_coluna_link(ws)
    indice_paginas = montar_indice_paginas(ws)

    for item in resultados:
        pagina = (item.get("pagina") or "Não identificado").strip()
        link = (item.get("url") or "").strip()
        page_id = (item.get("page_id") or "").strip()
        total = item.get("total_anuncios_ativos", 0)

        chave = criar_chave_registro(page_id, pagina, link)

        if chave in indice_paginas:
            linha = indice_paginas[chave]
        else:
            linha = ws.max_row + 1
            ws.cell(row=linha, column=1, value=pagina)
            indice_paginas[chave] = linha

        ws.cell(row=linha, column=1, value=pagina)
        ws.cell(row=linha, column=col_data, value=total)
        ws.cell(row=linha, column=col_link, value=link)

        # Limpar e aplicar destaque
        celula = ws.cell(row=linha, column=col_data)
        celula.fill = PatternFill(fill_type=None)
        celula.font = FONT_NORMAL

        col_anterior = encontrar_coluna_data_anterior(ws, col_data)
        if col_anterior is not None:
            atual = converter_para_numero(total)
            anterior = converter_para_numero(ws.cell(row=linha, column=col_anterior).value)

            if atual is not None and anterior is not None:
                if atual > anterior:
                    celula.fill = FILL_SUBIU
                    celula.font = FONT_SUBIU
                elif atual < anterior:
                    celula.fill = FILL_CAIU
                    celula.font = FONT_CAIU
                else:
                    celula.fill = FILL_IGUAL
                    celula.font = FONT_IGUAL

    garantir_link_na_ultima_coluna(ws)

    # Estilizar cabeçalho
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Largura das colunas
    ws.column_dimensions["A"].width = 40
    for col in range(2, ws.max_column + 1):
        letra = get_column_letter(col)
        titulo = ws.cell(row=1, column=col).value
        if titulo == "Link da biblioteca":
            ws.column_dimensions[letra].width = 95
        else:
            ws.column_dimensions[letra].width = 14

    wb.save(arquivo)


# ── Main ──

def main():
    caminho = sys.argv[1] if len(sys.argv) > 1 else ARQUIVO_EXCEL_PADRAO

    print(f"=== Monitor Meta Ads - Atualização Automática ===")
    print(f"Planilha: {caminho}")
    print()

    # 1. Ler links da planilha existente
    urls = ler_links_da_planilha(caminho)
    if not urls:
        print("Nenhum link encontrado na planilha. Nada a fazer.")
        sys.exit(0)

    print(f"Links encontrados: {len(urls)}")
    print()

    # 2. Abrir browser e consultar cada link
    resultados = []

    with sync_playwright() as p:
        print("Abrindo navegador...")
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ]
        )

        context = browser.new_context(
            viewport={"width": 1440, "height": 1200},
            locale="pt-BR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/133.0.0.0 Safari/537.36"
            ),
        )

        for idx, url in enumerate(urls, start=1):
            page_id = extrair_page_id_da_url(url)
            print(f"  [{idx}/{len(urls)}] Consultando page_id={page_id or '?'}...", end=" ", flush=True)

            page = context.new_page()
            try:
                resultado = extrair_com_js(page, url, page_id)
            finally:
                try:
                    page.close()
                except Exception:
                    pass

            resultados.append(resultado)

            total = resultado.get("total_anuncios_ativos", 0)
            nome = resultado.get("pagina", "?")
            status = resultado.get("status", "?")

            if status == "ok":
                print(f"OK  |  {nome}  |  {total} anúncios ativos")
            else:
                msg = resultado.get("mensagem", "")
                print(f"ERRO  |  {nome}  |  {msg}")

        browser.close()

    # 3. Salvar na planilha
    print()
    print("Salvando resultados na planilha...")
    salvar_resultados_no_excel(resultados, caminho)
    print(f"Planilha atualizada: {Path(caminho).resolve()}")
    print("Concluído!")


if __name__ == "__main__":
    main()
